import io
import os
import re
import openpyxl
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Font

MONTHS_RU_GEN = ['', 'января', 'февраля', 'марта', 'апреля', 'мая', 'июня',
                  'июля', 'августа', 'сентября', 'октября', 'ноября', 'декабря']
MONTHS_RU_NOM = ['', 'январь', 'февраль', 'март', 'апрель', 'май', 'июнь',
                  'июль', 'август', 'сентябрь', 'октябрь', 'ноябрь', 'декабрь']

TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), 'template.xlsx')

# Weights from official scoring table (used only for reference / external display)
WEIGHTS = {
    'order_higher':     7,   # п.30.1 — BU
    'order_academy':    4,   # п.30.2 — BV
    'conference':       5,   # п.24 — BI, доклад на конференции
    'software':         5,   # ПО → п.27.2 (BP), 1 балл вклада = 0.2 в ячейке
    'article_vak_rinc': 8,   # п.27.1 — BO
    'article_rinc':     5,   # п.27.2 — BP
    'article_closed':   5,   # п.27.3 — BQ
}

# Excel column indices (1-based) for criterion columns in the template
# Row 17 contains the weight coefficients per event; row 18+ are data rows.
# C(3)=п.1 ... BW(75)=п.31
_C = {
    'conference': 61,  # BI  п.24    Доклад на конференции, вес=5
    'vak_rinc': 67,  # BO  п.27.1  ВАК+РИНЦ, вес=8
    'rinc':     68,  # BP  п.27.2  РИНЦ, вес=5  (сюда же идут ПО)
    'closed':   69,  # BQ  п.27.3  Закрытое, вес=5
    'higher':   73,  # BU  п.30.1  вышестоящий, вес=7
    'academy':  74,  # BV  п.30.2  ВАС, вес=4
    'total':    76,  # BX  Итого (SUMPRODUCT)
    'notes':    77,  # BY  примечания
}


def _set(ws, row, col, value):
    """Set cell value, skipping read-only MergedCells."""
    cell = ws.cell(row=row, column=col)
    if not isinstance(cell, MergedCell):
        cell.value = value


def _inc(ws, row, col, delta=1):
    """Increment a numeric cell value."""
    cell = ws.cell(row=row, column=col)
    if not isinstance(cell, MergedCell):
        cell.value = (cell.value or 0) + delta


def generate_report_xlsx(profile: dict, report: dict) -> bytes:
    wb = openpyxl.load_workbook(TEMPLATE_PATH)
    ws = wb.active

    year = report.get('year', '')
    month = report.get('month', 1)
    month_gen = MONTHS_RU_GEN[month] if 1 <= month <= 12 else str(month)
    month_nom = MONTHS_RU_NOM[month] if 1 <= month <= 12 else str(month)

    # ── Update title (merged cell A2) ────────────────────────────────────────
    title_cell = ws['A2']
    if title_cell.value:
        # Replace month/year at end of title, e.g. "за июнь 2026 г."
        new_title = re.sub(
            r'за\s+\S+\s+\d{4}\s*г\.?',
            f'за {month_nom} {year} г.',
            title_cell.value,
            flags=re.IGNORECASE,
        )
        title_cell.value = new_title

    # ── Profile fields ───────────────────────────────────────────────────────
    last_name = profile.get('last_name', '')
    first_pat = profile.get('first_patronymic', '')   # "Александр Георгиевич"
    position  = profile.get('position', '')

    # Initials from "Александр Георгиевич" → "А.Г."
    parts = first_pat.split()
    initials = ''.join(p[0] + '.' for p in parts if p)   # "А.Г."

    # "Григоренко А.Г." for the data row
    fio_short = f"{last_name} {initials}".strip()

    # "А. Григоренко" for the signature block
    first_initial = (parts[0][0] + '. ') if parts else ''
    fio_sign = f"{first_initial}{last_name}".strip()

    # ── Data row ─────────────────────────────────────────────────────────────
    DATA_ROW = 18

    _set(ws, DATA_ROW, 1, 1)           # A — порядковый номер
    _set(ws, DATA_ROW, 2, fio_short)   # B — Григоренко А.Г.

    # ── Оперативные задания ───────────────────────────────────────────────────
    for o in report.get('orders', []):
        if o.get('level') == 'higher':
            _inc(ws, DATA_ROW, _C['higher'])
        else:
            _inc(ws, DATA_ROW, _C['academy'])

    # ── Доклады на конференциях (п.24, вес=5) ─────────────────────────────────
    # Каждый подтверждённый доклад = 1 в ячейке; SUMPRODUCT × 5 = баллы.
    for _c in report.get('conferences', []):
        _inc(ws, DATA_ROW, _C['conference'])

    # ── Статьи ───────────────────────────────────────────────────────────────
    # points_taken — доля автора. Вес колонки: vak_rinc=8, rinc=5, closed=5.
    # cell = points_taken / weight → SUMPRODUCT = points_taken.
    _ART_W = {'vak_rinc': 8, 'rinc': 5, 'closed': 5}
    _ART_C = {'vak_rinc': _C['vak_rinc'], 'rinc': _C['rinc'], 'closed': _C['closed']}
    art_pts: dict = {}
    for a in report.get('articles', []):
        atype = a.get('article_type', 'rinc')
        art_pts[atype] = art_pts.get(atype, 0.0) + a.get('points_taken', 0)
    for atype, pts in art_pts.items():
        if pts > 0 and atype in _ART_C:
            _set(ws, DATA_ROW, _ART_C[atype], round(pts / _ART_W[atype], 6))

    # ── ПО (свидетельство ФИПС) → п.27.2 BP (РИНЦ, вес=5) ──────────────────
    # ПО регистрируется в РИНЦ, поэтому добавляется в ту же колонку BP (п.27.2).
    # 1 балл авторского вклада = 0.2 в ячейке (points_taken / 5).
    # При весе 5: SUMPRODUCT даёт ровно points_taken итоговых баллов.
    total_po_pts = sum(sw.get('points_taken', 0) for sw in report.get('software', []))
    if total_po_pts > 0:
        existing_rinc = ws.cell(row=DATA_ROW, column=_C['rinc']).value or 0
        _set(ws, DATA_ROW, _C['rinc'], existing_rinc + round(total_po_pts / 5, 6))

    # ── Формула итого ─────────────────────────────────────────────────────────
    _set(ws, DATA_ROW, _C['total'],
         f'=SUMPRODUCT($C{DATA_ROW}:$BW{DATA_ROW},$C$17:$BW$17)')

    # ── Должность и подпись (строки 20–21) ───────────────────────────────────
    _set(ws, 20, 1, position)       # A20 — должность сотрудника (полностью)
    cell_pos = ws.cell(row=20, column=1)
    orig_font = cell_pos.font
    cell_pos.font = Font(
        name=orig_font.name or 'Times New Roman',
        size=orig_font.size or 14,
        bold=orig_font.bold,
        italic=True,
    )
    _set(ws, 21, 74, fio_sign)      # BV21 — А. Григоренко

    # ── Дата в подвале (строка 22) ────────────────────────────────────────────
    footer_cell = ws['A22']
    if footer_cell.value:
        footer_cell.value = re.sub(
            r'«.*?»\s+\S+\s+\d{4}\s*г\.?',
            f'«____» {month_nom} {year} г.',
            footer_cell.value,
        )
    else:
        footer_cell.value = f'«____» {month_nom} {year} г.'

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
